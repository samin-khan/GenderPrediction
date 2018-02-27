import pandas as pd
import numpy as np
from sklearn.model_selection import LeaveOneOut
from sklearn.linear_model import LogisticRegression
import operator
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from string import ascii_lowercase

df = pd.read_excel("MergedData.xlsx")


"""COMBINED-FEATURE ANALYSIS"""

""" Option to choose which features to drop"""
dropList = ["Name", "dataID", "Last Letter", "Last Two"]
dfLooTest = df.drop(dropList, axis=1)

# List to keep track of indices where error occurs
errorIndices = []
index = 0

# Timing LOOCV runtime
import time
t0 = time.time()
    
# Uses Leave One Out Cross Validation (LOOCV) to estimate how training model would do on predicting external data
loo = LeaveOneOut()
x = dfLooTest.drop(['Gender'], axis=1).as_matrix()
y = dfLooTest[['Gender']].as_matrix().ravel()

male_total = 0
correct_male = 0
female_total = 0
correct_female = 0

# Can optionally recreate a model to exclude the last training round
model = LogisticRegression(C=1e5)

for train_index, test_index in loo.split(x):
    print("TRAIN:", train_index, "TEST:", test_index)
    x_train, x_test = x[train_index], x[test_index]
    y_train, y_test = y[train_index], y[test_index]
    model.fit(x_train, y_train)
    prediction = model.predict(x_test)
    print("Prediction: " + str(prediction))
    print(x_train, x_test, y_train, y_test)
    print(y_test == [0])
    if y_test == [0]:
        male_total += 1
        correct_male += 1 if prediction == y_test else 0
    else:
        female_total += 1
        correct_female += 1 if prediction == y_test else 0
    if prediction != y_test:
        errorIndices.append(index)
    index += 1
overallAcc = (correct_male + correct_female) / (male_total + female_total)
maleAcc = correct_male / male_total
femaleAcc = correct_female / female_total

t1 = time.time()
runtime = t1 - t0

# Coefficients and accuracy of training model against itself using LOOCV
coefficients = {}
for i in range(len(dfLooTest.columns) - 1):
    coefficients[dfLooTest.columns[i + 1]] = model.coef_[0][i]

sorted_x = sorted(coefficients.items(), key=operator.itemgetter(1), reverse=True)
print("Runtime of compact model LOOCV: " + str(runtime))
print("Overall accuracy: " + str(overallAcc))
print("Male accuracy: " + str(maleAcc))
print("Female accuracy: " + str(femaleAcc))

# Stores dataframe containing weights from compact model coefficients
dfCompactWeights = pd.DataFrame(sorted_x)
excel_dir = 'Merged Data Compact Analysis.xlsx'
book = load_workbook(excel_dir)
with pd.ExcelWriter(excel_dir, engine='openpyxl') as writer:
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)    

    ## Your dataframe to append. 
    dfCompactWeights.to_excel(writer, 'Sheet1', index=False)

    writer.save()  
writer.close()


""" SINGLE FEATURE ANALYSIS """
# Timing LOOCV runtime
t0 = time.time()

singleFeatureAcc = [('Feature', 'Overall Accuracy', 'Male Accuracy', 'Female Accuracy')]

for currFeature in df:
    if currFeature in dropList:
        continue
    dfLooTest = df[['Gender', currFeature]]

    # Uses Leave One Out Cross Validation (LOOCV) to estimate how training model would do on predicting external data
    loo = LeaveOneOut()
    x = dfLooTest.drop(['Gender'], axis=1).as_matrix()
    y = dfLooTest[['Gender']].as_matrix().ravel()

    male_total = 0
    correct_male = 0
    female_total = 0
    correct_female = 0

    # Can optionally recreate a model to exclude the last training round
    model = LogisticRegression(C=1)

    for train_index, test_index in loo.split(x):
        print("TRAIN:", train_index, "TEST:", test_index)
        x_train, x_test = x[train_index], x[test_index]
        y_train, y_test = y[train_index], y[test_index]
        model.fit(x_train, y_train)
        prediction = model.predict(x_test)
        print("Prediction: " + str(prediction))
        print(x_train, x_test, y_train, y_test)
        print(y_test==[0])
        if y_test == [0]:
            male_total += 1
            correct_male += 1 if prediction == y_test else 0
        else:
            female_total += 1
            correct_female += 1 if prediction == y_test else 0
    overallAcc = (correct_male + correct_female) / (male_total + female_total)
    maleAcc = correct_male / male_total
    femaleAcc = correct_female / female_total

    singleFeatureAcc.append((currFeature, overallAcc, maleAcc, femaleAcc))

t1 = time.time()
singleFeatureRunTime = t1 - t0
print("Runtime of compact model single feature analysis: " + str(singleFeatureRunTime))

dfWrite = pd.DataFrame(singleFeatureAcc)

excel_dir = 'Merged Data Compact Analysis.xlsx'
book = load_workbook(excel_dir)
with pd.ExcelWriter(excel_dir, engine='openpyxl') as writer:
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)    

    ## Your dataframe to append. 
    dfWrite.to_excel(writer, 'Sheet2', index=False)

    writer.save()  
writer.close()
