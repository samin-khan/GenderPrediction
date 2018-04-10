import pandas as pd
import numpy as np
from sklearn.model_selection import LeaveOneOut
from sklearn.linear_model import LogisticRegression
import operator
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from string import ascii_uppercase
from string import ascii_lowercase
import time

# Reading in data and changing the gender of all names that are both male and female to neutral
dfRaw = pd.read_excel("RACodedNamesFeb1.xlsx")
dfRaw["is_unique"] = ~dfRaw['Name'].duplicated(keep=False)
dfRaw.loc[(dfRaw["is_unique"] == False), 'gender'] = 'N'

# Change keep -> True to keep neutral genders
dfAllFeatures = dfRaw.drop_duplicates(subset=['Name'], keep=False)

# dfHead turns male and females values to numeric binaries
dfAllFeatures = dfAllFeatures[dfAllFeatures.gender != 'N'] # Gets rid of neutrals
dfAllFeatures = dfAllFeatures.replace({'M':0}, regex=True)
dfAllFeatures = dfAllFeatures.replace({'F':1}, regex=True)

""" One-hot First/Last Single/Two Letter(s) """
# Create one-hots for all first/last single/double then do single feature analysis vs congregated
dfFirstLastLetters = dfAllFeatures[['gender', 'First Letter', 'First Two', 'Last Letter', 'Last Two']]
dfNonCompact = dfAllFeatures[['gender']]

# 1-hot First Letter
for c1 in ascii_uppercase:
    
    """ Use indicator to check if column contains any 1s. If not, remove column"""
    
    dfFirstLastLetters.insert(len(dfFirstLastLetters.columns), c1, 0)
    dfFirstLastLetters.loc[(dfFirstLastLetters['First Letter'] == c1), c1] = 1
    dfNonCompact['FirstLetter: ' + c1] = dfFirstLastLetters[c1]
    
    dfFirstLastLetters.insert(len(dfFirstLastLetters.columns), c1.lower(), 0)
    dfFirstLastLetters.loc[(dfFirstLastLetters['Last Letter'] == c1.lower()), c1.lower()] = 1
    dfNonCompact['LastLetter: ' + c1.lower()] = dfFirstLastLetters[c1.lower()]
    for c2 in ascii_lowercase:
        dfFirstLastLetters.insert(len(dfFirstLastLetters.columns), c1+c2, 0)
        dfFirstLastLetters.loc[(dfFirstLastLetters['First Two'] == c1+c2), c1+c2] = 1
        dfNonCompact['FirstTwo: ' + c1+c2] = dfFirstLastLetters[c1+c2]
        
        dfFirstLastLetters.insert(len(dfFirstLastLetters.columns), c1.lower()+c2, 0)
        dfFirstLastLetters.loc[(dfFirstLastLetters['Last Two'] == c1.lower()+c2), c1.lower()+c2] = 1
        dfNonCompact['LastTwo: ' + c1.lower()+c2] = dfFirstLastLetters[c1.lower()+c2]
        
"""LOOCV ANALYSIS"""
# Change parameter depending on target feature
dfLooTest = dfNonCompact

# Timing LOOCV runtime
t0 = time.time()
    
# Uses Leave One Out Cross Validation (LOOCV) to estimate how training model would do on predicting external data
loo = LeaveOneOut() 
x = dfLooTest.drop(['gender'], axis=1).as_matrix()
y = dfLooTest[['gender']].as_matrix().ravel()

male_total = 0
correct_male = 0
female_total = 0
correct_female = 0

# Can optionally recreate a model to exclude the last training round
model = LogisticRegression(C=1e5)

for train_index, test_index in loo.split(x):
    print("TRAIN:", train_index, "TEST:", test_index)
    x_train, x_test = x[train_index], x[test_index]
    y_train, y_test = y[train_index], y[test_index]
    model.fit(x_train, y_train)
    prediction = model.predict(x_test)
    print("Prediction: " + str(prediction))
    print(x_train, x_test, y_train, y_test)
    if y_test == [0]:
        male_total += 1
        correct_male += 1 if prediction == y_test else 0
    else:
        female_total += 1
        correct_female += 1 if prediction == y_test else 0
overallAcc = (correct_male + correct_female) / (male_total + female_total)
maleAcc = correct_male / male_total
femaleAcc = correct_female / female_total

t1 = time.time()
runtime = t1 - t0

# Coefficients and accuracy of training model against itself using LOOCV
coefficients = {}
for i in range(len(dfLooTest.columns) - 1):
    coefficients[dfLooTest.columns[i + 1]] = model.coef_[0][i]

sorted_x = sorted(coefficients.items(), key=operator.itemgetter(1), reverse=True)
print("Runtime of whole model: " + str(runtime))
print("Overall accuracy: " + str(overallAcc))
print("Male accuracy: " + str(maleAcc))
print("Female accuracy: " + str(femaleAcc))
print("\nCoefficients: \n")
print({'Total male': male_total, 'Correct male': correct_male, 'Total female': female_total, 'Correct female': correct_female})

# Storing dataframe containing model's feature weight coefficients
featureWeights = pd.DataFrame(sorted_x)
excel_dir = '1-hot First-Last Letters.xlsx'
book = load_workbook(excel_dir)
with pd.ExcelWriter(excel_dir, engine='openpyxl') as writer:
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)    

    ## Your dataframe to append. 
    featureWeights.to_excel(writer, 'Sheet1', index=False)

    writer.save()  
writer.close()

""" SINGLE FEATURE ANALYIS """
firstLetter = [f for f in dfNonCompact.columns if "FirstLetter" in f]
firstTwo = [f for f in dfNonCompact.columns if "FirstTwo" in f]
lastLetter = [f for f in dfNonCompact.columns if "LastLetter" in f]
lastTwo = [f for f in dfNonCompact.columns if "LastTwo" in f]

letterFeatures = {}
letterFeatures["FirstLetter"] = dfNonCompact[['gender'] + firstLetter]
letterFeatures["FirstTwo"] = dfNonCompact[['gender'] + firstTwo]
letterFeatures["LastLetter"] = dfNonCompact[['gender'] + lastLetter]
letterFeatures["LastTwo"] = dfNonCompact[['gender'] + lastTwo]

"""LOOCV LOOPED ANALYSIS"""

features = letterFeatures

oneHotAnalysis = [('One-hot vectors',), ('Feature', 'Overall Accuracy', 'Male Accuracy', 'Female Accuracy')]


# Timing LOOCV runtime
t0 = time.time()

for feature in features.keys():

    # Change parameter depending on target feature
    dfLooTest = features[feature]

    # Uses Leave One Out Cross Validation (LOOCV) to estimate how training model would do on predicting external data
    loo = LeaveOneOut()
    x = dfLooTest.drop(['gender'], axis=1).as_matrix()
    y = dfLooTest[['gender']].as_matrix().ravel()

    male_total = 0
    correct_male = 0
    female_total = 0
    correct_female = 0

    # Can optionally recreate a model to exclude the last training round
    model = LogisticRegression(C=1e5)

    for train_index, test_index in loo.split(x):
        print("TRAIN:", train_index, "TEST:", test_index)
        x_train, x_test = x[train_index], x[test_index]
        y_train, y_test = y[train_index], y[test_index]
        model.fit(x_train, y_train)
        prediction = model.predict(x_test)
        print("Prediction: " + str(prediction))
        print(x_train, x_test, y_train, y_test)
        if y_test == [0]:
            male_total += 1
            correct_male += 1 if prediction == y_test else 0
        else:
            female_total += 1
            correct_female += 1 if prediction == y_test else 0
    overallAcc = (correct_male + correct_female) / (male_total + female_total)
    maleAcc = correct_male / male_total
    femaleAcc = correct_female / female_total
    oneHotAnalysis.append((feature, overallAcc, maleAcc, femaleAcc))

t1 = time.time()
runtime = t1 - t0
print("Runtime of single feature analysis: " + str(runtime))

# Storing dataframe containing accuracies of each feature
oneHotAccuracies = pd.DataFrame(oneHotAnalysis)
excel_dir = '1-hot First-Last Letters.xlsx'
book = load_workbook(excel_dir)
with pd.ExcelWriter(excel_dir, engine='openpyxl') as writer:
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)    

    ## Your dataframe to append. 
    oneHotAccuracies.to_excel(writer, 'Sheet2', index=False)

    writer.save()  
writer.close()
